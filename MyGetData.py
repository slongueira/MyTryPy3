import os
import sys
import time
import shutil
import numpy as np
import pandas as pd
import pyqtgraph as pg
from datetime import datetime
from ctypes import byref, c_int32
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from PyQt5.QtWidgets import (QApplication, QPushButton, QVBoxLayout, QWidget,
                             QLabel, QSpinBox, QHBoxLayout, QFileDialog, QInputDialog)
from PyDAQmx.DAQmxConstants import (DAQmx_Val_RSE, DAQmx_Val_Volts, 
                                    DAQmx_Val_Rising, DAQmx_Val_ContSamps, 
                                    DAQmx_Val_GroupByScanNumber, DAQmx_Val_Acquired_Into_Buffer, 
                                    DAQmx_Val_GroupByChannel, DAQmx_Val_ChanForAllLines)
from PyQt5.QtCore import Qt, QObject, pyqtSignal, QThread, QTimer, pyqtSlot
from PyDAQmx import Task
from RaspberryInterface import RaspberryInterface
from MyMerger import Files_merge
from pyqtgraph.parametertree import Parameter, ParameterTree

# ---------------- CONFIG ----------------

CHANNELS = {
    "LinMot_Enable": "Dev1/ai0",
    "LinMot_Up_Down": "Dev1/ai1",
    "Voltage": "Dev1/ai2",
    "Current": "Dev1/ai3"
}

SAMPLE_RATE = 10000
SAMPLES_PER_CALLBACK = 100
CALLBACKS_PER_BUFFER = 500
BUFFER_SIZE = SAMPLES_PER_CALLBACK * CALLBACKS_PER_BUFFER

TimeWindowLength = 3  # seconds
PLOT_BUFFER_SIZE = ((SAMPLE_RATE * TimeWindowLength) // SAMPLES_PER_CALLBACK) * SAMPLES_PER_CALLBACK
refresh_rate = 10

moveLinMot = False

# ---------------- BUFFER PROCESSING THREAD ----------------
class BufferProcessor(QObject):
    process_buffer = pyqtSignal(object)

    def __init__(self, fs, parent=None):
        super().__init__(parent)
        self.fs = fs
        self.process_buffer.connect(self.save_data)
        self.timestamp = 0
        self.local_path = None

    @pyqtSlot(object)
    def save_data(self, data):
        if moveLinMot:
            t = np.arange(data.shape[0]) / self.fs + self.timestamp
            self.timestamp = t[-1] + (t[1] - t[0])
            df = pd.DataFrame({
                "Time (s)": t,
                "Voltage": data[:, 2],
                "Current": data[:, 3],
                "LINMOT_ENABLE": data[:, 0],
                "LINMOT_UP_DOWN": data[:, 1]
            })
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            df.to_pickle(f"{self.local_path}/DAQ_{timestamp}.pkl")
            print(f"[+] Saved {len(data)} samples")

# ---------------- DAQ TASK WITH CALLBACK ----------------
class DAQTask(Task):
    def __init__(self, plot_buffer, processor_signal, data_column_selector):
        super().__init__()

        self.plot_buffer = plot_buffer
        self.write_index = 0
        self.processor_signal = processor_signal
        self.data_column_selector = data_column_selector

        self.buffer1 = np.empty((BUFFER_SIZE, 4))
        self.buffer2 = np.empty((BUFFER_SIZE, 4))
        self.current_buffer = self.buffer1
        self.index = 0
        
        for channel in list(CHANNELS.values()):
            self.CreateAIVoltageChan(channel, "", DAQmx_Val_RSE, -10.0, 10.0, DAQmx_Val_Volts, None)
        
        self.CfgSampClkTiming("", SAMPLE_RATE, DAQmx_Val_Rising, DAQmx_Val_ContSamps, SAMPLES_PER_CALLBACK)
        self.AutoRegisterEveryNSamplesEvent(DAQmx_Val_Acquired_Into_Buffer, SAMPLES_PER_CALLBACK, 0)
        self.StartTask()

    def EveryNCallback(self):
        try:
            data = np.empty((SAMPLES_PER_CALLBACK, 4), dtype=np.float64)
            read = c_int32()
            self.ReadAnalogF64(SAMPLES_PER_CALLBACK, 10.0, DAQmx_Val_GroupByScanNumber, data, data.size, byref(read), None)
            
            data[:, 0] = np.where(data[:, 0] < 2, 0, 1)
            data[:, 1] = np.where(data[:, 1] < 2, 0, 1)
            data[:, 3] /= 499e3 
            
            self.plot_buffer[self.write_index:self.write_index + SAMPLES_PER_CALLBACK] = data[:, int(self.data_column_selector.value()[-1])]
            self.write_index = (self.write_index + SAMPLES_PER_CALLBACK) % self.plot_buffer.size
            
            self.current_buffer[self.index:self.index + SAMPLES_PER_CALLBACK, :] = data
            self.index += SAMPLES_PER_CALLBACK
            
            if self.index >= BUFFER_SIZE:
                full_buffer = self.current_buffer.copy()
                self.current_buffer = self.buffer1 if self.current_buffer is self.buffer2 else self.buffer2
                self.index = 0
                self.processor_signal.emit(full_buffer)
        
        except Exception as e:
            print(f"DAQ error in callback: {e}")
            
        return 0

# ---------------- INTERFACE AND PLOT  ----------------
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DAQ Viewer")
        self.layout = QVBoxLayout(self)
        
        # Request experiments directory
        print("Please select the experiment directory.")
        self.exp_dir = QFileDialog.getExistingDirectory(self, "Select Experiment Directory")
        if not self.exp_dir or not os.path.isdir(self.exp_dir):
            print("No directory selected. Exiting.")
            sys.exit(0)
        self.exp_dir = os.path.normpath(self.exp_dir)
            
        # Request TribuId
        print("Please enter TribuId.")
        self.tribu_id, ok = QInputDialog.getText(self, "Input", "Enter TribuId:")
        if not ok or not self.tribu_id:
            print("No TribuId entered. Exiting.")
            sys.exit(0)

        self.plot_buffer = np.zeros(PLOT_BUFFER_SIZE, dtype=float)

        self.plot_widget = pg.PlotWidget()
        self.curve = self.plot_widget.plot(self.plot_buffer, pen='y')
        
        # Raspberry Pi SSH connection parameters
        hostname = "192.168.100.200"
        port = 22
        username = "TENG"
        password = "raspberry"
        self.remote_path = "/var/opt/codesys/PlcLogic/FTP_Folder"

        self.raspberry = RaspberryInterface(hostname=hostname, port=port, username=username, password=password)
        self.thread_raspberry = QThread()
        self.raspberry.moveToThread(self.thread_raspberry)
        self.thread_raspberry.start()
        self.raspberry.execute.emit(lambda: self.raspberry.connect())

        self.button = QPushButton("START LinMot")
        self.button.clicked.connect(self.toggle_linmot)
        
        # Timer UI elements
        self.timer_label = QLabel("Duration (s):")
        self.timer_spinbox = QSpinBox()
        self.timer_spinbox.setRange(1, 86400)  # 1 sec to 24 hours
        self.timer_spinbox.setValue(30)  # Default value
        self.countdown_display = QLabel("Remaining time: -")
        self.countdown_display.setAlignment(Qt.AlignRight)
        duration = QHBoxLayout()
        duration.addWidget(self.timer_label)
        duration.addWidget(self.timer_spinbox)
        
        # Signal selector
        self.param_group = Parameter.create(name='Signal', type='list', value=CHANNELS['Voltage'], limits=CHANNELS)
        self.tree = ParameterTree()
        self.tree.setParameters(self.param_group, showTop=True)
        self.layout.addWidget(self.tree)
        
        self.layout.addWidget(self.button)
        self.layout.addWidget(self.plot_widget)
        self.layout.addLayout(duration)
        self.layout.addWidget(self.countdown_display)
        
        # Timer setup
        self.measurement_timer = QTimer()
        self.measurement_timer.timeout.connect(self.update_countdown)
        self.remaining_seconds = 0
        self.should_save_data = False
        
        # Buffer processor and thread
        self.processor = BufferProcessor(SAMPLE_RATE)
        self.thread = QThread()
        self.processor.moveToThread(self.thread)
        self.thread.start()

        self.task = DAQTask(self.plot_buffer, self.processor.process_buffer, self.param_group)
        
        # Digital IO tasks
        self.DO_task_LinMotTrigger = DigitalOutputTask(line="Dev1/port0/line7")
        self.DO_task_LinMotTrigger.StartTask()
        
        self.DO_task_RelayLine0 = DigitalOutputTask(line="Dev1/port0/line5")
        self.DO_task_RelayLine0.StartTask()

        self.DO_task_PrepareRaspberry = DigitalOutputTask(line="Dev1/port0/line6")
        self.DO_task_PrepareRaspberry.StartTask()

        self.DI_task_Raspberry_status_0 = DigitalInputTask(line="Dev1/port1/line0")
        self.DI_task_Raspberry_status_0.StartTask()

        self.DI_task_Raspberry_status_1 = DigitalInputTask(line="Dev1/port1/line1")
        self.DI_task_Raspberry_status_1.StartTask()
        
        # Plot update timer
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_plot)
        self.timer.start(refresh_rate)
    
    def update_countdown(self):
        if self.remaining_seconds > 0 and moveLinMot:
            self.remaining_seconds -= 1
            self.countdown_display.setText(f"Remaining time: {self.remaining_seconds} s")
        else:
            self.measurement_timer.stop()
            self.countdown_display.setText("Remaining time: -")
            self.should_save_data = True
            self.toggle_linmot()
    
    def update_plot(self):
        display_data = np.concatenate((
            self.plot_buffer[self.task.write_index:],
            self.plot_buffer[:self.task.write_index]
        ))
        self.curve.setData(display_data)
    
    def closeEvent(self, event):
        # Cleanup DAQ tasks
        self.task.StopTask()
        self.task.ClearTask()

        self.DO_task_LinMotTrigger.set_line(0)
        self.DO_task_LinMotTrigger.StopTask()
        self.DO_task_LinMotTrigger.ClearTask()
        
        self.DO_task_RelayLine0.set_line(0)
        self.DO_task_RelayLine0.StopTask()
        self.DO_task_RelayLine0.ClearTask()

        self.DO_task_PrepareRaspberry.set_line(0)
        self.DO_task_PrepareRaspberry.StopTask()
        self.DO_task_PrepareRaspberry.ClearTask()

        self.DI_task_Raspberry_status_0.StopTask()
        self.DI_task_Raspberry_status_0.ClearTask()

        self.DI_task_Raspberry_status_1.StopTask()
        self.DI_task_Raspberry_status_1.ClearTask()

        self.thread.quit()
        self.thread.wait()
        
        self.thread_raspberry.quit()
        self.thread_raspberry.wait()
        
        if moveLinMot and os.path.isdir(self.processor.local_path):
            shutil.rmtree(self.processor.local_path)
            print(f"Temporary folder {self.processor.local_path} deleted on exit.")
        
        event.accept()

    def toggle_linmot(self):
        global moveLinMot
        
        if moveLinMot:
            # STOP measurement
            self.measurement_timer.stop()
            self.countdown_display.setText("Remaining time: -")
            
            self.DO_task_LinMotTrigger.set_line(0)
            self.DO_task_PrepareRaspberry.set_line(0)
            self.DO_task_RelayLine0.set_line(0)

            if self.task.index != 0:
                data = self.task.current_buffer[:self.task.index]
                self.task.processor_signal.emit(data)
                self.task.index = 0

            loop_counter = 0
            while loop_counter < 10000:
                status_bit_0 = self.DI_task_Raspberry_status_0.read_line()
                status_bit_1 = self.DI_task_Raspberry_status_1.read_line()
                if status_bit_0 == 0 and status_bit_1 == 0:
                    break
                loop_counter += 1

            if loop_counter >= 10000:
                print("\033[91mError loop counter overflow, raspberry is not responding\033[0m")
                return

            self.raspberry.execute.emit(lambda: self.raspberry.download_folder(self.remote_path, local_path=self.processor.local_path))
            self.raspberry.execute.emit(lambda: self.raspberry.remove_files_with_extension(self.remote_path))

            time.sleep(1)
            if self.should_save_data:
                self.motor_file, self.daq_file = Files_merge(folder_path=self.processor.local_path, exp_id=self.exp_id)
                self.add_experiment_row()
            else:
                print("Experiment interrupted.")
            
            shutil.rmtree(self.processor.local_path)

        else:
            # START measurement
            print("\nPlease enter RloadId")
            self.rload_id, ok = QInputDialog.getText(self, "Input", "Enter RloadId:")
            if not ok or not self.rload_id:
                print("No RloadId entered. Operation canceled.")
                return
            
            self.date_now = datetime.now().strftime("%d%m%Y_%H%M%S")
            self.exp_id = f"{self.date_now}-{self.tribu_id}-{self.rload_id}"

            self.DO_task_PrepareRaspberry.set_line(1)
            
            loop_counter = 0
            while loop_counter < 10000:
                status_bit_0 = self.DI_task_Raspberry_status_0.read_line()
                status_bit_1 = self.DI_task_Raspberry_status_1.read_line()

                if status_bit_0 == 0 and status_bit_1 == 0:
                    loop_counter += 1
                elif status_bit_0 == 1 and status_bit_1 == 0:
                    break
                elif status_bit_0 == 0 and status_bit_1 == 1:
                    self.DO_task_PrepareRaspberry.set_line(0)
                    self.raspberry.execute.emit(lambda: self.raspberry.reset_codesys())
                    print("\033[91mError, impossible to prepare raspberry to record, check codesys invalid license error. Resetting Codesys, please wait... \033[0m")
                    return
                else:
                    self.DO_task_PrepareRaspberry.set_line(0)
                    self.raspberry.execute.emit(lambda: self.raspberry.reset_codesys())
                    print("\033[91mError, EtherCAT bus is not working, resetting Codesys, please wait...\033[0m")
                    return

            if loop_counter >= 10000:
                self.DO_task_PrepareRaspberry.set_line(0)
                print("\033[91mError loop counter overflow, raspberry is not responding\033[0m")
                return
            
            self.DO_task_RelayLine0.set_line(1)
            
            os.makedirs(os.path.join(self.exp_dir, "RawData"), exist_ok=True)
            self.processor.local_path = os.path.join(self.exp_dir, "RawData", self.exp_id)
            os.makedirs(self.processor.local_path, exist_ok=True)
            
            self.processor.timestamp = 0
            
            self.remaining_seconds = self.timer_spinbox.value()
            self.countdown_display.setText(f"Remaining time: {self.remaining_seconds} s")
            self.measurement_timer.start(1000)  # 1 sec
            self.should_save_data = False
            
            self.task.index = 0
            self.DO_task_LinMotTrigger.set_line(1)
            
        moveLinMot = not moveLinMot
        self.button.setText("STOP LinMot" if moveLinMot else "START LinMot")
        
    def add_experiment_row(self):
        """Add a new row to ExpsDescription.xlsx with the experiment data if it doesn't already exist."""
        excel_path = os.path.join(self.exp_dir, "ExpsDescription.xlsx")
        
        if os.path.isfile(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
            if not ws.tables:
                print(f"Could not save the experiment row because the Excel file {excel_path} has no tables.")
                return
        else:
            print(f"Could not save the experiment row because the Excel file {excel_path} was not found.")
            return

        table_name = list(ws.tables.keys())[0]
        table = ws.tables[table_name]
        start_cell, end_cell = table.ref.split(":")
        start_col = "".join(filter(str.isalpha, start_cell))
        start_row = int("".join(filter(str.isdigit, start_cell)))
        end_col = "".join(filter(str.isalpha, end_cell))
        end_row = int("".join(filter(str.isdigit, end_cell)))

        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        # Create the new row to be inserted
        new_row = [
            self.exp_id,
            self.tribu_id,
            self.date_now,
            self.daq_file,
            self.motor_file,
            "",
            self.rload_id
        ] + [""] * 23  # Empty columns to fill with blank spaces

        # Find the first empty row
        for row_idx in range(start_row, end_row + 1):
            row_cells = ws[row_idx]
            if all(cell.value in (None, "") for cell in row_cells[start_col_idx - 1:end_col_idx]):
                first_empty_row = row_idx
                break
        else:
            first_empty_row = end_row + 1

        # Insert the new row
        for i, value in enumerate(new_row, start=start_col_idx):
            ws.cell(row=first_empty_row, column=i, value=value)

        # Update the table range if needed
        new_end_row = max(end_row, first_empty_row)
        if new_end_row != end_row:
            new_ref = f"{start_col}{start_row}:{end_col}{new_end_row}"
            table.ref = new_ref

        # Adjust column widths (optional)
        for col_idx in range(start_col_idx, end_col_idx + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in range(start_row, new_end_row + 1):  # Include new row
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(excel_path)

# ---------------- DIGITAL IO TASKS ----------------
class DigitalOutputTask(Task):
    def __init__(self, line):
        super().__init__()
        self.CreateDOChan(line, "", DAQmx_Val_ChanForAllLines)
        self.set_line(0)

    def set_line(self, value):
        data = np.array([value], dtype=np.uint8)
        self.WriteDigitalLines(1, 1, 10.0, DAQmx_Val_GroupByChannel, data, None, None)

class DigitalInputTask(Task):
    def __init__(self, line):
        super().__init__()
        self.CreateDIChan(line, "", DAQmx_Val_ChanForAllLines)

    def read_line(self):
        data = np.zeros(1, dtype=np.uint8)
        read = c_int32()
        self.ReadDigitalLines(1, 10.0, 0, data, 1, read, None, None)
        return data[0]

# ---------------- MAIN ----------------
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


